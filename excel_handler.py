import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter  # Add this import
from openpyxl.workbook import Workbook  # Also missing in your original code
import os
import logging
from typing import Optional, Dict, List


class ExcelHandler:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.export_folder = 'exports'
    
    def read_student_data(self, filepath: str) -> Optional[pd.DataFrame]:
        """
        Read student data from Excel file.
        Expected columns: Roll Number, Name, Class, Section, Gender
        """
        try:
            # Try reading the Excel file
            df = pd.read_excel(filepath)
            
            # Normalize column names (handle case variations and spaces)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            # Define expected column mappings
            column_mappings = {
                'roll_number': ['roll_number', 'roll', 'rollno', 'roll_no', 'student_id'],
                'name': ['name', 'student_name', 'full_name'],
                'class': ['class', 'class_name', 'std', 'standard', 'grade'],
                'section': ['section', 'sec', 'division'],
                'gender': ['gender', 'sex', 'm/f']
            }
            
            # Map columns
            mapped_columns = {}
            for expected_col, possible_names in column_mappings.items():
                for possible_name in possible_names:
                    if possible_name in df.columns:
                        mapped_columns[expected_col] = possible_name
                        break
            
            # Check if all required columns are found
            required_columns = ['roll_number', 'name', 'class', 'section', 'gender']
            missing_columns = []
            
            for req_col in required_columns:
                if req_col not in mapped_columns:
                    missing_columns.append(req_col)
            
            if missing_columns:
                self.logger.error(f"Missing columns: {missing_columns}")
                return None
            
            # Create new DataFrame with standardized column names
            result_df = pd.DataFrame()
            for standard_name, original_name in mapped_columns.items():
                result_df[standard_name] = df[original_name]
            
            # Clean and validate data
            result_df = self._clean_student_data(result_df)
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"Error reading Excel file: {str(e)}")
            return None
    
    def _clean_student_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and validate student data.
        """
        try:
            # Remove rows with missing essential data
            df = df.dropna(subset=['roll_number', 'name', 'class', 'section'])
            
            # Convert to string and strip whitespace
            for col in ['roll_number', 'name', 'class', 'section', 'gender']:
                df[col] = df[col].astype(str).str.strip()
            
            # Standardize gender values
            gender_mapping = {
                'male': 'Male', 'm': 'Male', 'boy': 'Male',
                'female': 'Female', 'f': 'Female', 'girl': 'Female'
            }
            df['gender'] = df['gender'].str.lower().map(gender_mapping).fillna(df['gender'])
            
            # Remove duplicate roll numbers
            df = df.drop_duplicates(subset=['roll_number'], keep='first')
            
            # Sort by class, section, and roll number
            df = df.sort_values(['class', 'section', 'roll_number'])
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error cleaning student data: {str(e)}")
            return df
    
    def export_room_seating(self, room_id: str, room_info: Dict, room_seating: Dict) -> Optional[str]:
        """
        Export room seating arrangement to Excel file.
        """
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = f"Room {room_id} Seating"
            
            # Set up styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Add header information
            ws['A1'] = f"Examination Seating Arrangement - Room {room_id}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"Room Capacity: {room_info.get('rows', 0)} rows × {room_info.get('columns', 0)} columns"
            ws.merge_cells('A2:D2')
            
            # Add table headers
            headers = ['Desk Number', 'Seat A', 'Seat B', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Add seating data
            row_num = 5
            desk_count = 0
            
            # Sort desks by row and column
            sorted_desks = sorted(room_seating.items(), 
                                key=lambda x: (x[1]['row'], x[1]['col']))
            
            for desk_id, desk_info in sorted_desks:
                desk_count += 1
                
                # Format student information
                seat_a_info = self._format_student_info(desk_info.get('seat_a'))
                seat_b_info = self._format_student_info(desk_info.get('seat_b'))
                
                # Determine notes
                notes = ""
                if not seat_a_info and not seat_b_info:
                    notes = "Empty"
                elif not seat_a_info or not seat_b_info:
                    notes = "Half occupied"
                
                # Add row data
                row_data = [desk_id, seat_a_info, seat_b_info, notes]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Color coding for easier reading
                    if notes == "Empty":
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif notes == "Half occupied":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                row_num += 1
            
            # Add summary
            ws.cell(row=row_num + 1, column=1, value="Summary:")
            ws.cell(row=row_num + 1, column=1).font = Font(bold=True)
            
            total_capacity = desk_count * 2
            occupied_seats = sum(1 for desk in room_seating.values() 
                               for seat in ['seat_a', 'seat_b'] 
                               if desk.get(seat))
            
            ws.cell(row=row_num + 2, column=1, value=f"Total Desks: {desk_count}")
            ws.cell(row=row_num + 3, column=1, value=f"Total Capacity: {total_capacity}")
            ws.cell(row=row_num + 4, column=1, value=f"Occupied Seats: {occupied_seats}")
            ws.cell(row=row_num + 5, column=1, value=f"Empty Seats: {total_capacity - occupied_seats}")
            
            # Auto-adjust column widths
            for col_idx in range(1, 5):  # Columns A through D
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, row_num + 6):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"{room_id}_seating.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)
            
            self.logger.info(f"Exported seating plan to {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error exporting room seating: {str(e)}")
            return None
    
    def _format_student_info(self, student: Optional[Dict]) -> str:
        """
        Format student information for display.
        Format: RollNumber - Name (Class Section Gender)
        """
        if not student:
            return ""
        
        try:
            roll = student.get('roll_number', 'N/A')
            name = student.get('name', 'N/A')
            class_name = student.get('class', 'N/A')
            section = student.get('section', 'N/A')
            gender = student.get('gender', 'N/A')
            
            return f"{roll} - {name} ({class_name} {section} {gender})"
            
        except Exception as e:
            self.logger.error(f"Error formatting student info: {str(e)}")
            return "Error"
    
    def export_room_grid_layout(self, room_id: str, room_info: Dict, room_seating: Dict) -> Optional[str]:
        """
        Export room seating plan in row/column grid format for visual representation.
        """
        try:
            from datetime import datetime
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{room_id} Grid Layout"
            
            # Room configuration
            rows = room_info.get('rows', 8)
            columns = room_info.get('columns', 6)
            extra_desks = room_info.get('extra_desks', 0)
            
            # Title
            ws.merge_cells('A1:' + get_column_letter(columns * 3) + '1')
            title_cell = ws.cell(row=1, column=1, value=f"Room {room_id} - Grid Layout")
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            current_row = 3
            ws.cell(row=current_row, column=1, value="Row/Col").font = Font(bold=True)
            
            # Column headers
            for col in range(columns):
                ws.cell(row=current_row, column=(col * 3) + 2, value=f"Desk {col + 1}").font = Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=(col * 3) + 2, 
                              end_row=current_row, end_column=(col * 3) + 3)
            
            current_row += 1
            
            # Sub-headers for seats
            for col in range(columns):
                ws.cell(row=current_row, column=(col * 3) + 2, value="Seat A").font = Font(size=9)
                ws.cell(row=current_row, column=(col * 3) + 3, value="Seat B").font = Font(size=9)
            
            current_row += 1
            
            # Grid layout
            for row in range(rows):
                ws.cell(row=current_row, column=1, value=f"Row {row + 1}").font = Font(bold=True)
                
                for col in range(columns):
                    desk_id = f"R{row + 1}C{col + 1}"
                    desk_data = room_seating.get(desk_id, {'seat_a': None, 'seat_b': None})
                    
                    # Seat A
                    seat_a_cell = ws.cell(row=current_row, column=(col * 3) + 2)
                    seat_b_cell = ws.cell(row=current_row, column=(col * 3) + 3)
                    
                    seat_a_info = self._format_student_short(desk_data.get('seat_a'))
                    seat_b_info = self._format_student_short(desk_data.get('seat_b'))
                    
                    seat_a_cell.value = seat_a_info or "Empty"
                    seat_b_cell.value = seat_b_info or "Empty"
                    
                    # Color coding
                    if not seat_a_info:
                        seat_a_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    else:
                        seat_a_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    
                    if not seat_b_info:
                        seat_b_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    else:
                        seat_b_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    
                    # Borders
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    seat_a_cell.border = thin_border
                    seat_b_cell.border = thin_border
                    
                current_row += 1
            
            # Extra desks section
            if extra_desks > 0:
                current_row += 2
                ws.cell(row=current_row, column=1, value="Extra Desks:").font = Font(bold=True)
                current_row += 1
                
                for extra in range(extra_desks):
                    desk_id = f"E{extra + 1}"
                    desk_data = room_seating.get(desk_id, {'seat_a': None, 'seat_b': None})
                    
                    ws.cell(row=current_row, column=1, value=f"Extra {extra + 1}")
                    ws.cell(row=current_row, column=2, value=self._format_student_short(desk_data.get('seat_a')) or "Empty")
                    ws.cell(row=current_row, column=3, value=self._format_student_short(desk_data.get('seat_b')) or "Empty")
                    current_row += 1
            
            # Summary
            current_row += 2
            ws.cell(row=current_row, column=1, value="Room Summary:").font = Font(bold=True, size=12)
            current_row += 1
            
            total_desks = rows * columns + extra_desks
            total_capacity = total_desks * 2
            occupied_seats = sum(1 for desk in room_seating.values() 
                               for seat in ['seat_a', 'seat_b'] 
                               if desk.get(seat))
            
            ws.cell(row=current_row, column=1, value=f"Total Desks: {total_desks}")
            ws.cell(row=current_row + 1, column=1, value=f"Total Capacity: {total_capacity}")
            ws.cell(row=current_row + 2, column=1, value=f"Occupied Seats: {occupied_seats}")
            ws.cell(row=current_row + 3, column=1, value=f"Utilization: {round((occupied_seats/total_capacity)*100, 1)}%")
            
            # Auto-adjust column widths
            for col_idx in range(1, (columns * 3) + 2):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15
            
            # Save file
            filename = f"{room_id}_grid_layout.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)
            
            self.logger.info(f"Exported grid layout to {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error exporting grid layout: {str(e)}")
            return None

    def _format_student_short(self, student: Optional[Dict]) -> str:
        """Format student info in short format for grid view."""
        if not student:
            return ""
        
        try:
            roll = student.get('roll_number', 'N/A')
            name = student.get('name', 'N/A')
            # Truncate name if too long
            if len(name) > 12:
                name = name[:10] + ".."
            return f"{roll}\n{name}"
        except Exception:
            return "Error"

    def export_all_rooms_zip(self, seating_plan: Dict, rooms_info: List[Dict]) -> Optional[str]:
        """
        Export all room seating plans as separate Excel files and create a ZIP.
        """
        try:
            import zipfile
            from datetime import datetime
            
            exported_files = []
            
            # Export each room (both formats)
            for room_id, room_seating in seating_plan.items():
                room_info = next((r for r in rooms_info if r['room_id'] == room_id), {})
                
                # Export standard format
                filepath1 = self.export_room_seating(room_id, room_info, room_seating)
                if filepath1:
                    exported_files.append(filepath1)
                
                # Export grid format
                filepath2 = self.export_room_grid_layout(room_id, room_info, room_seating)
                if filepath2:
                    exported_files.append(filepath2)
            
            # Create summary workbook
            summary_file = self.create_summary_workbook(seating_plan, rooms_info)
            if summary_file:
                exported_files.append(summary_file)
            
            if not exported_files:
                return None
            
            # Create ZIP file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            zip_filename = f"seating_plan_complete_{timestamp}.zip"
            zip_filepath = os.path.join(self.export_folder, zip_filename)
            
            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for file_path in exported_files:
                    if os.path.exists(file_path):
                        # Add file to ZIP with just the filename (no path)
                        zip_file.write(file_path, os.path.basename(file_path))
            
            # Clean up individual files
            for file_path in exported_files:
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception:
                    pass  # Ignore cleanup errors
            
            self.logger.info(f"Created complete seating plan ZIP: {zip_filepath}")
            return zip_filepath
            
        except Exception as e:
            self.logger.error(f"Error creating ZIP export: {str(e)}")
            return None

    def create_summary_workbook(self, seating_plan: Dict, rooms_info: List[Dict]) -> Optional[str]:
        """
        Create a summary workbook with overview of all rooms.
        """
        try:
            from datetime import datetime
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Seating Plan Summary"
            
            # Title
            ws.merge_cells('A1:H1')
            title_cell = ws.cell(row=1, column=1, value="PM SHRI JNV GB NAGAR - Exam Seating Plan Summary")
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Generation info
            ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws.cell(row=2, column=1).font = Font(size=10, italic=True)
            
            # Headers
            current_row = 4
            headers = ['Room ID', 'Layout', 'Capacity', 'Occupied', 'Empty', 'Utilization %', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
            current_row += 1
            
            # Room summaries
            total_capacity = 0
            total_occupied = 0
            
            for room_id, room_seating in seating_plan.items():
                room_info = next((r for r in rooms_info if r['room_id'] == room_id), {})
                
                rows = room_info.get('rows', 0)
                columns = room_info.get('columns', 0)
                extra_desks = room_info.get('extra_desks', 0)
                capacity = room_info.get('capacity', 0)
                
                occupied = sum(1 for desk in room_seating.values() 
                             for seat in ['seat_a', 'seat_b'] 
                             if desk.get(seat))
                empty = capacity - occupied
                utilization = round((occupied / capacity * 100), 1) if capacity > 0 else 0
                
                total_capacity += capacity
                total_occupied += occupied
                
                # Status
                if utilization >= 90:
                    status = "Full"
                    status_color = "FFE6E6"
                elif utilization >= 70:
                    status = "Good"
                    status_color = "FFF2CC"
                else:
                    status = "Low"
                    status_color = "E6F3FF"
                
                # Row data
                row_data = [
                    room_id,
                    f"{rows}×{columns}+{extra_desks}",
                    capacity,
                    occupied,
                    empty,
                    f"{utilization}%",
                    status
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    if col == 7:  # Status column
                        cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                
                current_row += 1
            
            # Totals row
            current_row += 1
            ws.cell(row=current_row, column=1, value="TOTALS").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=total_capacity).font = Font(bold=True)
            ws.cell(row=current_row, column=4, value=total_occupied).font = Font(bold=True)
            ws.cell(row=current_row, column=5, value=total_capacity - total_occupied).font = Font(bold=True)
            overall_util = round((total_occupied / total_capacity * 100), 1) if total_capacity > 0 else 0
            ws.cell(row=current_row, column=6, value=f"{overall_util}%").font = Font(bold=True)
            
            # Statistics section
            current_row += 3
            ws.cell(row=current_row, column=1, value="Overall Statistics:").font = Font(bold=True, size=12)
            current_row += 1
            
            stats = [
                f"Total Rooms: {len(seating_plan)}",
                f"Total Capacity: {total_capacity} students",
                f"Students Assigned: {total_occupied}",
                f"Empty Seats: {total_capacity - total_occupied}",
                f"Overall Utilization: {overall_util}%"
            ]
            
            for stat in stats:
                ws.cell(row=current_row, column=1, value=stat)
                current_row += 1
            
            # Auto-adjust column widths
            for col_idx in range(1, 8):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15
            
            # Save file
            filename = f"seating_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)
            
            self.logger.info(f"Created summary workbook: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating summary workbook: {str(e)}")
            return None
            
            with zipfile.ZipFile(zip_filepath, 'w') as zipf:
                for file_path in exported_files:
                    zipf.write(file_path, os.path.basename(file_path))
            
            return zip_filepath
            
        except Exception as e:
            self.logger.error(f"Error creating ZIP export: {str(e)}")
            return None
